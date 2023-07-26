import openpyxl
from docx import Document

def read_excel_and_write_to_word(excel_file, word_file):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(excel_file)
    # Get the active sheet
    sheet = workbook.active

    # Create a new Word document
    document = Document()

    # Create a table in the Word document
    table = document.add_table(rows=1, cols=sheet.max_column)

    # Iterate over each row in the Excel sheet
    for row in sheet.iter_rows():
        # Add a new row to the table
        table_row = table.add_row().cells
        # Iterate over each cell in the row
        for i, cell in enumerate(row):
            # Get the cell value
            cell_value = cell.value
            # Add the cell value to the corresponding cell in the table
            table_row[i].text = str(cell_value)

    # Save the Word document
    document.save(word_file)

    print("Data successfully copied to Word file.")


# Usage example
excel_file_path = "example.xlsx"
word_file_path = "output.docx"

read_excel_and_write_to_word(excel_file_path, word_file_path)
